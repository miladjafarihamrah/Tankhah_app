from django.shortcuts import render, redirect
from django.contrib.auth import login, authenticate
from django.conf import settings
from django.http import HttpResponse
from django.contrib import messages
from django.shortcuts import get_object_or_404

from .forms import SignUpForm, LoginForm
from .forms import MissionForm
from .forms import ExpenseForm
from .forms import KhodroForm
from .forms import ReportForm
from django.db.models import Sum



from .models import Mission
from .models import Expense  
from .models import Balance
from .models import Khodro
from .models import TransactionHistory


import jdatetime
from django.http import JsonResponse
import os
from datetime import datetime
from openpyxl import load_workbook, Workbook
from django.contrib.auth.decorators import login_required
from .forms import UserUpdateForm
import json

# استفاده از requests اگر موجود باشد، در غیر این صورت از urllib
try:
    import requests
    HAS_REQUESTS = True
except ImportError:
    HAS_REQUESTS = False
    try:
        from urllib.request import urlopen, Request
        from urllib.error import URLError
    except ImportError:
        pass
# دیکشنری برای تبدیل نام ماه‌های شمسی به انگلیسی
PERSIAN_MONTH_TO_ENGLISH = {
    'فروردین': 'Farvardin',
    'اردیبهشت': 'Ordibehesht',
    'خرداد': 'Khordad',
    'تیر': 'Tir',
    'مرداد': 'Mordad',
    'شهریور': 'Shahrivar',
    'مهر': 'Mehr',
    'آبان': 'Aban',
    'آذر': 'Azar',
    'دی': 'Dey',
    'بهمن': 'Bahman',
    'اسفند': 'Esfand'
}

# دیکشنری برای تبدیل نام روزهای هفته به فارسی
PERSIAN_WEEKDAYS = {
    0: 'شنبه',
    1: 'یکشنبه',
    2: 'دوشنبه',
    3: 'سه‌شنبه',
    4: 'چهارشنبه',
    5: 'پنج‌شنبه',
    6: 'جمعه'
}

# تابع تبدیل اعداد انگلیسی به فارسی
def to_persian_digits(text):
    persian_digits = ['۰', '۱', '۲', '۳', '۴', '۵', '۶', '۷', '۸', '۹']
    english_digits = ['0', '1', '2', '3', '4', '5', '6', '7', '8', '9']
    for i, digit in enumerate(english_digits):
        text = text.replace(digit, persian_digits[i])
    return text


def signup(request):
    if request.method == 'POST':
        form = SignUpForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            return redirect('home')
    else:
        form = SignUpForm()
    return render(request, 'signup.html', {'form': form})

def user_login(request):
    if request.method == 'POST':
        form = LoginForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                return redirect('home')
    else:
        form = LoginForm()
    return render(request, 'login.html', {'form': form})
#ویرایش نام کاربری ،پسورود ،نام ونام خانوادگی
@login_required
def update_profile(request):
    if request.method == 'POST':
        form = UserUpdateForm(request.POST, instance=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, "اطلاعات شما با موفقیت به‌روزرسانی شد.")
            return redirect('home')
    else:
        form = UserUpdateForm(instance=request.user)

    return render(request, 'update_profile.html', {'form': form})

#صفحه ابزار
@login_required
def tools(request):
    if not request.user.is_authenticated:
        return redirect('login')
    return render(request, 'tools.html')

#صفحه انتخاب نرم افزار
@login_required
def select_software(request):
    if not request.user.is_authenticated:
        return redirect('login')
    return render(request, 'select_software.html')

#صفحه منوال دستگاه ها
@login_required
def device_manual(request):
    if not request.user.is_authenticated:
        return redirect('login')
    return render(request, 'device_manual.html')

#صفحه قیمت طلا و دلار
@login_required
def gold_price(request):
    if not request.user.is_authenticated:
        return redirect('login')
    
    # دریافت قیمت‌ها از API
    gold_price_18k = None
    silver_price = None
    dollar_price = None
    error_message = None
    
    def get_price_from_api():
        """تابع برای دریافت قیمت‌ها از API جایگزین"""
        gold_18k = None
        silver = None
        dollar = None
        
        if not HAS_REQUESTS:
            return None, None, None
        
        # روش 1: استفاده از API ساده و مستقیم
        try:
            # دریافت قیمت طلا 18 عیار
            gold_url = 'https://call.tgju.org/ajax.json?p=geram18'
            gold_resp = requests.get(gold_url, timeout=8, headers={'User-Agent': 'Mozilla/5.0'})
            if gold_resp.status_code == 200:
                gold_data = gold_resp.json()
                if isinstance(gold_data, dict):
                    # بررسی ساختارهای مختلف
                    if 'p' in gold_data:
                        gold_18k = int(float(gold_data['p']))
                    elif 'price' in gold_data:
                        gold_18k = int(float(gold_data['price']))
                    elif 'value' in gold_data:
                        gold_18k = int(float(gold_data['value']))
        except Exception as e:
            pass
        
        try:
            # دریافت قیمت نقره
            silver_url = 'https://call.tgju.org/ajax.json?p=geram999'
            silver_resp = requests.get(silver_url, timeout=8, headers={'User-Agent': 'Mozilla/5.0'})
            if silver_resp.status_code == 200:
                silver_data = silver_resp.json()
                if isinstance(silver_data, dict):
                    if 'p' in silver_data:
                        silver = int(float(silver_data['p']))
                    elif 'price' in silver_data:
                        silver = int(float(silver_data['price']))
                    elif 'value' in silver_data:
                        silver = int(float(silver_data['value']))
        except Exception as e:
            pass
        
        try:
            # دریافت قیمت دلار
            dollar_url = 'https://call.tgju.org/ajax.json?p=usd'
            dollar_resp = requests.get(dollar_url, timeout=8, headers={'User-Agent': 'Mozilla/5.0'})
            if dollar_resp.status_code == 200:
                dollar_data = dollar_resp.json()
                if isinstance(dollar_data, dict):
                    if 'p' in dollar_data:
                        dollar = int(float(dollar_data['p']))
                    elif 'price' in dollar_data:
                        dollar = int(float(dollar_data['price']))
                    elif 'value' in dollar_data:
                        dollar = int(float(dollar_data['value']))
        except Exception as e:
            pass
        
        # اگر قیمت‌ها دریافت شدند، برگردان (حتی اگر یکی از آن‌ها None باشد)
        if gold_18k or silver or dollar:
            return gold_18k, silver, dollar
        
        # روش 2: استفاده از API navasan
        try:
            api_url = 'https://api.navasan.tech/latest/?item=geram18,geram999,usd'
            response = requests.get(api_url, timeout=8, headers={'User-Agent': 'Mozilla/5.0'})
            if response.status_code == 200:
                data = response.json()
                if isinstance(data, dict):
                    if not gold_18k and 'geram18' in data:
                        gold_val = data['geram18']
                        if isinstance(gold_val, dict):
                            gold_18k = int(float(gold_val.get('value', gold_val.get('price', 0))))
                        else:
                            gold_18k = int(float(gold_val))
                    
                    if not silver and 'geram999' in data:
                        silver_val = data['geram999']
                        if isinstance(silver_val, dict):
                            silver = int(float(silver_val.get('value', silver_val.get('price', 0))))
                        else:
                            silver = int(float(silver_val))
                    
                    if not dollar and 'usd' in data:
                        dollar_val = data['usd']
                        if isinstance(dollar_val, dict):
                            dollar = int(float(dollar_val.get('value', dollar_val.get('price', 0))))
                        else:
                            dollar = int(float(dollar_val))
        except Exception as e:
            pass
        
        # روش 3: استفاده از API tgju با ساختار متفاوت
        try:
            # دریافت همه قیمت‌ها از یک endpoint
            all_url = 'https://api.tgju.org/v1/data/sanarate/v1'
            all_resp = requests.get(all_url, timeout=8, headers={'User-Agent': 'Mozilla/5.0'})
            if all_resp.status_code == 200:
                all_data = all_resp.json()
                if isinstance(all_data, dict) and 'data' in all_data:
                    prices_list = all_data['data']
                    if isinstance(prices_list, list):
                        for item in prices_list:
                            if isinstance(item, dict):
                                key = str(item.get('key', '')).lower()
                                price_val = item.get('p', item.get('price', 0))
                                
                                if not gold_18k and ('geram18' in key or 'gold18' in key):
                                    try:
                                        gold_18k = int(float(price_val))
                                    except:
                                        pass
                                
                                if not silver and ('geram999' in key or 'silver999' in key):
                                    try:
                                        silver = int(float(price_val))
                                    except:
                                        pass
                                
                                if not dollar and 'usd' in key:
                                    try:
                                        dollar = int(float(price_val))
                                    except:
                                        pass
        except Exception as e:
            pass
        
        return gold_18k, silver, dollar
    
    try:
        # استفاده از API جایگزین برای دریافت قیمت‌ها
        if HAS_REQUESTS:
            # دریافت قیمت‌ها از API جایگزین
            gold_18k, silver, dollar = get_price_from_api()
            
            if gold_18k and gold_18k > 0:
                gold_price_18k = gold_18k
            if silver and silver > 0:
                silver_price = silver
            if dollar and dollar > 0:
                dollar_price = dollar
            
            # اگر هنوز قیمت‌ها پیدا نشد، از endpoint کلی استفاده کن
            if not gold_price_18k or not silver_price or not dollar_price:
                try:
                    all_prices_url = 'https://api.tgju.org/v1/data/sanarate/v1'
                    all_response = requests.get(all_prices_url, timeout=10, headers={'User-Agent': 'Mozilla/5.0'})
                    if all_response.status_code == 200:
                        all_data = all_response.json()
                        if isinstance(all_data, dict) and 'data' in all_data:
                            prices_list = all_data['data']
                            if isinstance(prices_list, list):
                                for item in prices_list:
                                    if isinstance(item, dict):
                                        key = str(item.get('key', '')).lower()
                                        price = item.get('p', item.get('price', 0))
                                        
                                        if not gold_price_18k and ('geram18' in key or 'gold18' in key):
                                            try:
                                                gold_price_18k = int(float(price))
                                            except:
                                                pass
                                        
                                        if not silver_price and ('geram999' in key or 'silver999' in key):
                                            try:
                                                silver_price = int(float(price))
                                            except:
                                                pass
                                        
                                        if not dollar_price and 'usd' in key:
                                            try:
                                                dollar_price = int(float(price))
                                            except:
                                                pass
                except:
                    pass
        else:
            # استفاده از urllib
            try:
                from urllib.request import urlopen, Request
                from urllib.error import URLError
                
                # دریافت قیمت طلا
                try:
                    gold_url = 'https://api.tgju.org/v1/data/sanarate/geram18'
                    req = Request(gold_url, headers={'User-Agent': 'Mozilla/5.0'})
                    with urlopen(req, timeout=10) as response:
                        gold_data = json.loads(response.read().decode())
                        if isinstance(gold_data, dict) and 'data' in gold_data:
                            price = gold_data['data'].get('p', gold_data['data'].get('price', 0))
                            if price:
                                gold_price_18k = int(float(price))
                except:
                    pass
                
                # دریافت قیمت نقره
                try:
                    silver_url = 'https://api.tgju.org/v1/data/sanarate/geram999'
                    req = Request(silver_url, headers={'User-Agent': 'Mozilla/5.0'})
                    with urlopen(req, timeout=10) as response:
                        silver_data = json.loads(response.read().decode())
                        if isinstance(silver_data, dict) and 'data' in silver_data:
                            price = silver_data['data'].get('p', silver_data['data'].get('price', 0))
                            if price:
                                silver_price = int(float(price))
                except:
                    pass
                
                # دریافت قیمت دلار
                try:
                    dollar_url = 'https://api.tgju.org/v1/data/sanarate/usd'
                    req = Request(dollar_url, headers={'User-Agent': 'Mozilla/5.0'})
                    with urlopen(req, timeout=10) as response:
                        dollar_data = json.loads(response.read().decode())
                        if isinstance(dollar_data, dict) and 'data' in dollar_data:
                            price = dollar_data['data'].get('p', dollar_data['data'].get('price', 0))
                            if price:
                                dollar_price = int(float(price))
                except:
                    pass
            except Exception as e:
                pass
        
        # بررسی وضعیت دریافت قیمت‌ها
        if not gold_price_18k and not silver_price and not dollar_price:
            error_message = "امکان دریافت قیمت‌ها از سرور وجود ندارد. لطفاً اتصال اینترنت خود را بررسی کنید و بعداً تلاش کنید."
        elif not gold_price_18k or not silver_price or not dollar_price:
            # اگر برخی قیمت‌ها دریافت شدند، پیام جزئی نمایش بده
            missing = []
            if not gold_price_18k:
                missing.append("طلا")
            if not silver_price:
                missing.append("نقره")
            if not dollar_price:
                missing.append("دلار")
            error_message = f"برخی قیمت‌ها دریافت نشد: {', '.join(missing)}"
            
    except Exception as e:
        error_message = f"خطا در اتصال به سرور: {str(e)}"
    
    # تبدیل قیمت‌ها به اعداد فارسی برای نمایش با جداسازی سه‌رقمی
    def format_persian_number(num):
        if num:
            # جداسازی سه‌رقمی
            num_str = f"{num:,}"
            # تبدیل به فارسی
            return to_persian_digits(num_str)
        return None
    
    gold_price_18k_persian = format_persian_number(gold_price_18k)
    silver_price_persian = format_persian_number(silver_price)
    dollar_price_persian = format_persian_number(dollar_price)
    
    return render(request, 'gold_price.html', {
        'gold_price_18k': gold_price_18k,
        'gold_price_18k_persian': gold_price_18k_persian,
        'silver_price': silver_price,
        'silver_price_persian': silver_price_persian,
        'dollar_price': dollar_price,
        'dollar_price_persian': dollar_price_persian,
        'error_message': error_message,
    })

#صفحه اصلی برنامه
def home(request):
    if not request.user.is_authenticated:
        return redirect('login')
    
    # دریافت تاریخ جاری به صورت شمسی
    current_jalali = jdatetime.datetime.now()
    today = current_jalali.strftime('%Y/%m/%d')
    # تبدیل اعداد تاریخ به فارسی
    today = to_persian_digits(today)
    
    # دریافت نام روز هفته
    weekday_name = PERSIAN_WEEKDAYS[current_jalali.weekday()]
    current_year = current_jalali.year
    current_month = current_jalali.month

# فیلتر ماموریت‌های ماه و سال جاری
    total = Mission.objects.filter(
    user=request.user,
    date__startswith=f"{current_year}/{current_month:02d}"
    ).aggregate(Sum('mission_units'))['mission_units__sum'] or 0

    total_mission_units = int(total) if float(total).is_integer() else total


    full_name = f"{request.user.first_name} {request.user.last_name}"
                
    # دریافت مانده حساب
    try:
        balance = Balance.objects.get(user=request.user).amount
    except Balance.DoesNotExist:
        # اگر مانده حساب وجود نداشت، یک مانده حساب جدید ایجاد کنید
        balance = 4000000  # مقدار پیش‌فرض
        Balance.objects.create(user=request.user, amount=balance)

    return render(request, 'home.html', {
        'full_name': full_name,
        'today': today,
        'weekday_name': weekday_name,
        'mission_count': total_mission_units,  # تعداد ماموریت‌های ماه جاری
        'balance': balance,  # ارسال مانده حساب به تمپلیت
    })

import os
from openpyxl import Workbook, load_workbook

def convert_persian_numbers_to_english(text):
    """تبدیل اعداد فارسی به انگلیسی"""
    persian_numbers = "۰۱۲۳۴۵۶۷۸۹"
    english_numbers = "0123456789"
    translation_table = str.maketrans(persian_numbers, english_numbers)
    return text.translate(translation_table)

def add_mission(request):
    if not request.user.is_authenticated:
        return redirect('login')

    if request.method == 'POST':
        form = MissionForm(request.POST)
        if form.is_valid():
            mission = form.save(commit=False)
            mission.user = request.user

            # تبدیل اعداد فارسی به انگلیسی در تاریخ
            mission.date = convert_persian_numbers_to_english(mission.date)

            # مقدار تنظیم‌شده را اعمال کن
            mission_type_received = request.POST.get('mission_type', 'normal')
            mission.mission_type = mission_type_received
            mission.mission_units = {
                'normal': 1.0,
                'half': 0.5,
                'holiday': 2.0
            }.get(mission.mission_type, 1.0)  

            # بررسی تکراری نبودن تاریخ ماموریت برای همین کاربر
            existing_mission = Mission.objects.filter(user=request.user, date=mission.date).exists()
            if existing_mission:
                messages.error(request, 'خطا: تاریخ ماموریت تکراری است.')
                return render(request, 'add_mission.html', {'form': form})

            mission.save()
            return redirect('home')
        else:
            messages.error(request, 'خطا در ثبت ماموریت. لطفاً دوباره تلاش کنید.')
    else:
        form = MissionForm()

    return render(request, 'add_mission.html', {'form': form})
    
def edit_mission(request):
    year = request.GET.get('year')
    month = request.GET.get('month')
    
    if not year or not month:
        return render(request, 'error.html', {'message': 'سال و ماه معتبر نیست.'})

    # فیلتر کردن ماموریت‌ها بر اساس سال و ماه
    missions = Mission.objects.filter(user=request.user)
    missions = [mission for mission in missions 
               if mission.date.split('/')[0] == year 
               and mission.date.split('/')[1] == month]

    return render(request, 'edit_mission.html', {
        'missions': missions, 
        'month': month,
        'year': year
    })

def delete_mission(request):
    if not request.user.is_authenticated:
        return JsonResponse({'status': 'error', 'message': 'احراز هویت لازم است.'})
    
    date = request.GET.get('date')  # دریافت تاریخ از درخواست GET
    if not date:
        return JsonResponse({'status': 'error', 'message': 'تاریخ نامعتبر است.'})

    # حذف ماموریت از دیتابیس
    mission = get_object_or_404(Mission, date=date, user=request.user)
    mission.delete()

    return JsonResponse({'status': 'success', 'message': 'ماموریت با موفقیت حذف شد.'})


def add_expense(request):
    if not request.user.is_authenticated:
        return redirect('login')

    if request.method == 'POST':
        form = ExpenseForm(request.POST)
        if form.is_valid():
            expense = form.save(commit=False)
            expense.user = request.user

            # تبدیل اعداد فارسی به انگلیسی در تاریخ
            expense.date = convert_persian_numbers_to_english(expense.date)
            
            # تبدیل مبلغ به عدد صحیح
            try:
                expense.amount = int(expense.amount)
            except ValueError:
                messages.error(request, 'مبلغ وارد شده معتبر نیست.')
                return render(request, 'add_expense.html', {'form': form})

            # کاهش مانده حساب
            try:
                balance = Balance.objects.get(user=request.user)
                balance.amount -= expense.amount
                balance.save()
            except Balance.DoesNotExist:
                messages.error(request, 'خطا: حسابی برای این کاربر ثبت نشده است.')
                return render(request, 'add_expense.html', {'form': form})

            expense.save()
            return redirect('home')
        else:
            messages.error(request, 'خطا در ثبت تنخواه. لطفاً دوباره تلاش کنید.')
    else:
        form = ExpenseForm()

    return render(request, 'add_expense.html', {'form': form})

# ویرایش هزینه‌ها
def edit_expense(request):
    if not request.user.is_authenticated:
        return redirect('login')
    
    year = request.GET.get('year')
    month = request.GET.get('month')
    
    if not year or not month:
        return render(request, 'error.html', {'message': 'سال و ماه معتبر نیست.'})

    # فیلتر کردن هزینه‌ها بر اساس ماه
    expenses = Expense.objects.filter(user=request.user)
    expenses = [expense for expense in expenses 
               if expense.date.split('/')[0] == year 
               and expense.date.split('/')[1] == month]

    return render(request, 'edit_expense.html', {
        'expenses': expenses,
        'month': month,
        'year': year
          })



def delete_expense(request):
    if not request.user.is_authenticated:
        return JsonResponse({'status': 'error', 'message': 'احراز هویت لازم است.'})
    
    expense_id = request.GET.get('id')  # دریافت id به جای date

    if not expense_id:
        return JsonResponse({'status': 'error', 'message': 'شناسه نامعتبر است.'})

    # حذف Expense بر اساس ID
    expense = get_object_or_404(Expense, id=expense_id, user=request.user)
    amount = expense.amount  # مقدار هزینه برای به‌روزرسانی مانده حساب
    expense.delete()

    # افزایش مانده حساب
    balance = Balance.objects.get(user=request.user)
    balance.amount += amount
    balance.save()

    return JsonResponse({'status': 'success', 'message': 'تنخواه با موفقیت حذف شد.'})

from django.shortcuts import render, redirect
from django.contrib import messages
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from datetime import datetime
from .forms import ReportForm
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.workbook import Workbook
from .models import Mission, Expense, Khodro

def generate_report(request):
    if not request.user.is_authenticated:
        return redirect('login')
    
    if request.method == 'POST':
        print("***********************************")

        form = ReportForm(request.POST)
        if form.is_valid():
            report_type = request.POST.get('report_type')
            REPORT_TYPES = ('mission', 'expense', 'khodro', 'mission_pdf', 'expense_pdf', 'khodro_pdf')
            if report_type not in REPORT_TYPES:
                messages.error(request, 'نوع گزارش نامعتبر است.')
                return redirect('home')
            
            # اگر نوع گزارش PDF است، مستقیماً view generate_pdf_report را فراخوانی کن
            if report_type.endswith('_pdf'):
                # فراخوانی مستقیم تابع generate_pdf_report
                return generate_pdf_report(request)
            
            # برای گزارش‌های سالیانه (غیر PDF)
            try:
                year = int(form.cleaned_data['year'])
                if year < 1400 or year > 1450:
                    messages.error(request, 'سال وارد شده نامعتبر است.')
                    return redirect('home')
            except (ValueError, KeyError):
                messages.error(request, 'خطا در پردازش داده‌ها.')
                return redirect('home')

            report_config = {
                'mission': {
                    'model': Mission,
                    'output_name': f"annual_mission_{request.user.username}_{year}.xlsx",
                    'header': ['ردیف', 'تاریخ', 'کارخانه']
                },
                'expense': {
                    'model': Expense,
                    'output_name': f"annual_expense_{request.user.username}_{year}.xlsx",
                    'header': ['ردیف', 'تاریخ', 'توضیحات', 'مبلغ(ریال)', 'کارخانه']
                },
                'khodro': {
                    'model': Khodro,
                    'output_name': f"annual_khodro_{request.user.username}_{year}.xlsx",
                    'header': ['ردیف', 'تاریخ', 'کیلومتر', 'شرح سرویس', 'مبلغ(ریال)']
                }
            }

            config = report_config[report_type]
            model = config['model']
            output_name = config['output_name']
            header = config['header']

            # فیلتر داده‌ها و مرتب‌سازی بر اساس تاریخ
            data = model.objects.filter(user=request.user)
            data_list = []

            for item in data:
                try:
                    item_year, item_month, item_day = map(int, item.date.split('/'))
                    if item_year == year:
                        data_list.append((item_year, item_month, item_day, item))
                except:
                    continue  # رد کردن رکوردهای تاریخ نامعتبر

            # مرتب‌سازی بر اساس سال، ماه، روز
            data_list.sort(key=lambda x: (x[0], x[1], x[2]))

            filtered_data = []

            if report_type == 'mission':
                mission_count = 0
                for _, _, _, item in data_list:
                    filtered_data.append([item.date, item.factory])
                    mission_count += 1
                filtered_data.append(["", f"جمع ماموریت‌ها: {mission_count}"])
            
            else:
                total_amount = 0
                monthly_total = 0
                previous_month = None

                for _, item_month, _, item in data_list:
                    if previous_month is None:
                        previous_month = item_month

                    # اگر ماه تغییر کرد، جمع ماه قبل را اضافه کن
                    if item_month != previous_month:
                        filtered_data.append(["", "", f"جمع ماه {previous_month}:", convert_to_persian_numbers(monthly_total)])
                        monthly_total = 0
                        previous_month = item_month

                    if report_type == 'expense':
                        filtered_data.append([
                            item.date,
                            item.description,
                            convert_to_persian_numbers(item.amount),
                            item.factory
                        ])
                        monthly_total += item.amount
                        total_amount += item.amount
                    elif report_type == 'khodro':
                        filtered_data.append([
                            item.date,
                            item.description,
                            convert_to_persian_numbers(item.kilometer),
                            convert_to_persian_numbers(item.amount)
                        ])
                        monthly_total += item.amount
                        total_amount += item.amount

                # جمع ماه آخر
                filtered_data.append(["", "", f"جمع ماه {previous_month}:", convert_to_persian_numbers(monthly_total)])
                # جمع کل
                filtered_data.append(["", "", "جمع کل(ریال):", convert_to_persian_numbers(total_amount)])

            # ایجاد فایل اکسل
            wb = Workbook()
            ws = wb.active
            ws.sheet_view.rightToLeft = True
            ws.append(header)

            for index, row in enumerate(filtered_data, 1):
                ws.append([index] + row)

            # استایل‌دهی هدر
            header_font = Font(bold=True, color="FFFFFF", size=14, name="B Nazanin")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            border = Border(left=Side(style='thin'), right=Side(style='thin'),
                            top=Side(style='thin'), bottom=Side(style='thin'))
            alignment = Alignment(horizontal='center', vertical='center')

            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = alignment

            # استایل‌دهی داده‌ها
            data_font = Font(size=12, name="B Nazanin")
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(header)+1):
                for cell in row:
                    cell.font = data_font
                    cell.border = border
                    cell.alignment = alignment

            # فرمت اعداد برای مالی
            if report_type in ['expense', 'khodro']:
                amount_column = 4
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=amount_column, max_col=amount_column):
                    for cell in row:
                        try:
                            cell.value = int(str(cell.value).replace(',', ''))
                            cell.number_format = '#,##0'
                        except:
                            pass

            # تنظیم عرض ستون‌ها به صورت خودکار
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                    except:
                        pass
                ws.column_dimensions[column].width = max_length + 2

            # رنگ پس‌زمینه ماه‌ها متناوب برای expense و khodro
            if report_type in ['expense', 'khodro']:
                month_colors = ["DDDDDD", "FFFFFF"]  # روشن و تیره
                previous_month = None
                color_index = 0
                row_offset = 2  # شروع از ردیف 2 چون ردیف اول هدر است
                data_idx = 0  # برای دسترسی به data_list

                for row_idx in range(2, ws.max_row + 1):
                    cell_value = ws.cell(row=row_idx, column=2).value
                    if cell_value and '/' in str(cell_value):
                        # استخراج ماه از تاریخ
                        item_month = int(str(cell_value).split('/')[1])
                        if previous_month != item_month:
                            color_index = 1 - color_index
                            previous_month = item_month

                    # اعمال رنگ به کل ردیف
                    for col_idx in range(1, len(header)+2):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.fill = PatternFill(start_color=month_colors[color_index],
                                                end_color=month_colors[color_index],
                                                fill_type='solid')

            # ارسال فایل به کاربر
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{output_name}"'
            wb.save(response)
            return response

    else:
        form = ReportForm()

    return render(request, 'report.html', {'form': form})






def update_balance(request):
    if not request.user.is_authenticated:
        return redirect('login')
    
    if request.method == 'POST':
        amount = int(request.POST.get('amount', 0))
        action = request.POST.get('action')  # 'increase' یا 'decrease'

        if amount <= 0:
            messages.error(request, 'مبلغ وارد شده نامعتبر است.')
            return redirect('home')

        balance = Balance.objects.get(user=request.user)
        if action == 'increase':
            balance.amount += amount
        elif action == 'decrease':
            balance.amount -= amount
        balance.save()

        # ذخیره در تاریخچه
        TransactionHistory.objects.create(
            user=request.user,
            amount=amount if action == 'increase' else -amount,  # مثبت/منفی
            action=action
        )

        messages.success(request, f'مانده حساب با موفقیت {"افزایش" if action == "increase" else "کاهش"} یافت.')
        return redirect('home')
    else:
        return redirect('home')
    

from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import json
from .models import Mission  # مطمئن شو مدل درست ایمپورت شده

@csrf_exempt
def update_mission_factory(request, mission_id):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            factory_name = data.get('factory')

            mission = Mission.objects.get(id=mission_id)
            mission.factory = factory_name
            mission.save()

            return JsonResponse({'status': 'success'})
        except Mission.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'مأموریت پیدا نشد'})
    return JsonResponse({'status': 'error', 'message': 'درخواست نامعتبر'})

# views.py

from django.shortcuts import render
from django.http import JsonResponse
from .models import Expense, Balance  # فرض بر این است که مدل Expense و Balance موجود است

def edit_expense_details(request):
    expense_id = request.GET.get('id')
    factory = request.GET.get('factory')
    amount = request.GET.get('amount')
    description = request.GET.get('description')

    print(f"Expense ID: {expense_id}")  # برای دیباگ

    # بررسی اینکه مبلغ به عدد تبدیل شود
    try:
        amount = float(amount)  # تبدیل مبلغ به نوع عددی (float)
    except ValueError:
        return JsonResponse({'status': 'error', 'message': 'مقدار مبلغ معتبر نیست.'})

    try:
        # پیدا کردن رکورد مطابق با ID
        expense = Expense.objects.get(id=expense_id)

        # ذخیره مبلغ قدیمی برای به‌روزرسانی مانده حساب
        old_amount = expense.amount

        # به‌روزرسانی اطلاعات هزینه
        expense.factory = factory
        expense.amount = amount
        expense.description = description
        expense.save()  # ذخیره تغییرات در پایگاه داده

        # به‌روزرسانی مانده حساب
        balance = Balance.objects.get(user=request.user)
        
        # کم کردن مبلغ قدیمی و اضافه کردن مبلغ جدید
        balance.amount += (old_amount - amount)
        balance.save()
        

        return JsonResponse({'status': 'success'})
    except Expense.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'تنخواه پیدا نشد!'})
    except Balance.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'موجودی کاربر یافت نشد!'})

def hazineh_khodro(request):
    if not request.user.is_authenticated:
        return redirect('login')
    
    if request.method == 'POST':
        form = KhodroForm(request.POST)
        if form.is_valid():
            khodro = form.save(commit=False)
            khodro.user = request.user

            # تبدیل اعداد فارسی به انگلیسی در تاریخ
            khodro.date = convert_persian_numbers_to_english(khodro.date)
            
            # تبدیل مبلغ به عدد صحیح
            try:
                khodro.amount = int(khodro.amount)
            except ValueError:
                messages.error(request, 'مبلغ وارد شده معتبر نیست.')
                return render(request, 'hazineh_khodro.html', {'form': form})

            # کاهش مانده حساب
            balance = Balance.objects.get(user=request.user)
            balance.amount -= khodro.amount
            balance.save()

            khodro.save()
            return redirect('home')
        else:
            messages.error(request, 'خطا در ثبت تنخواه. لطفاً دوباره تلاش کنید.')
    else:
        form = KhodroForm()
    return render(request, 'hazineh_khodro.html', {'form': form})
# ویرایش هزینه‌ها
def edit_khodro(request):
    if not request.user.is_authenticated:
        return redirect('login')
    
    year = request.GET.get('year')
    month = request.GET.get('month')
    
    
    if not year or not month :
        return render(request, 'error.html', {'message': 'سال وماه  معتبر نیست.'})

    # فیلتر کردن هزینه‌ها بر اساس ماه
    khodros = Khodro.objects.filter(user=request.user)
    khodros = [khodro for khodro in khodros 
               if khodro.date.split('/')[0] == year
               and khodro.date.split('/')[1] == month]

    return render(request, 'edit_khodro.html', {
        'khodros': khodros,
        'month': month,
        'year': year,
          })


def delete_khodro(request):
    if not request.user.is_authenticated:
        return JsonResponse({'status': 'error', 'message': 'احراز هویت لازم است.'})
    
    khodro_id = request.GET.get('id')  # دریافت id به جای date

    if not khodro_id:
        return JsonResponse({'status': 'error', 'message': 'شناسه نامعتبر است.'})

    # حذف khodro بر اساس ID
    khodro = get_object_or_404(Khodro, id=khodro_id, user=request.user)
    amount = khodro.amount  # مقدار هزینه برای به‌روزرسانی مانده حساب
    khodro.delete()

    # افزایش مانده حساب
    balance = Balance.objects.get(user=request.user)
    balance.amount += amount
    balance.save()

    return JsonResponse({'status': 'success', 'message': 'هزینه خودرو با موفقیت حذف شد.'})

def edit_khodro_details(request):
    khodro_id = request.GET.get('id')
    kilometer = request.GET.get('kilometer')
    amount = request.GET.get('amount')
    description = request.GET.get('description')

    print(f"Khodro ID: {khodro_id}")  # برای دیباگ

    # بررسی اینکه مبلغ به عدد تبدیل شود
    try:
        amount = float(amount)  # تبدیل مبلغ به نوع عددی (float)
    except ValueError:
        return JsonResponse({'status': 'error', 'message': 'مقدار مبلغ معتبر نیست.'})

    try:
        # پیدا کردن رکورد مطابق با ID
        khodro = Khodro.objects.get(id=khodro_id)

        # ذخیره مبلغ قدیمی برای به‌روزرسانی مانده حساب
        old_amount = khodro.amount

        # به‌روزرسانی اطلاعات هزینه
        khodro.kilometer = kilometer
        khodro.amount = amount
        khodro.description = description
        khodro.save()  # ذخیره تغییرات در پایگاه داده

        # به‌روزرسانی مانده حساب
        balance = Balance.objects.get(user=request.user)
        
        # کم کردن مبلغ قدیمی و اضافه کردن مبلغ جدید
        balance.amount += (old_amount - amount)
        balance.save()

        return JsonResponse({'status': 'success'})
    except Khodro.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'تنخواه پیدا نشد!'})
    except Balance.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'موجودی کاربر یافت نشد!'})
MISSION_TYPE_DISPLAY = {
    'normal': '',
    'half': 'اصالت',
    'holiday': 'تعطیل'
}

# در ویو یا تمپلیت:
mission_type_display = MISSION_TYPE_DISPLAY.get(Mission.mission_type,'')

from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from arabic_reshaper import reshape
from bidi.algorithm import get_display
import os
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import urllib.request
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from django.conf import settings

# تنظیم مسیر فونت
FONT_PATH = os.path.join(settings.BASE_DIR, 'static', 'fonts', 'Vazir.ttf')

def generate_pdf_report(request):
    print("########################################################")

    if not request.user.is_authenticated:
        return redirect('login')
    
    if request.method == 'POST':
        print("########################################################")
        form = ReportForm(request.POST)
        if form.is_valid():
            try:
                month = int(form.cleaned_data['month'])
                year = int(form.cleaned_data['year'])
                
                if not (1 <= month <= 12):
                    messages.error(request, 'ماه وارد شده نامعتبر است.')
                    return redirect('home')
                
                if year < 1400 or year > 1450:
                    messages.error(request, 'سال وارد شده نامعتبر است.')
                    return redirect('home')

            except (ValueError, KeyError):
                messages.error(request, 'خطا در پردازش داده‌ها.')
                return redirect('home')

            report_type = request.POST.get('report_type')
            REPORT_TYPES = ('mission_pdf', 'expense_pdf', 'khodro_pdf')

            if report_type not in REPORT_TYPES:
                messages.error(request, 'نوع گزارش نامعتبر است.')
                return redirect('home')

            model_type = report_type.replace('_pdf', '')
            
            report_config = {
                'mission': {
                    'model': Mission,
                    'output_name': f"m_{request.user.username}_{year}_{month}.pdf",
                    'title': 'گزارش ماموریت‌ها',
                    'headers': ['کارخانه', 'تاریخ', 'ردیف']
                },
                'expense': {
                    'model': Expense,
                    'output_name': f"t_{request.user.username}_{year}_{month}.pdf",
                    'title': 'گزارش تنخواه‌ها',
                    'headers': ['کارخانه', 'مبلغ(ریال)', 'توضیحات', 'تاریخ', 'ردیف']
                },
                'khodro': {
                    'model': Khodro,
                    'output_name': f"k_{request.user.username}_{year}_{month}.pdf",
                    'title': 'گزارش هزینه‌های خودرو',
                    'headers': ['مبلغ(ریال)', 'شرح سرویس', 'کیلومتر', 'تاریخ', 'ردیف']
                }
            }

            config = report_config[model_type]
            model = config['model']
            output_name = config['output_name']
            title = config['title']
            headers = config['headers']

            # دریافت داده‌ها
            data = model.objects.filter(user=request.user)
            filtered_data = []
            total_amount = 0

            if model_type == 'mission':
                normal_missions = []
                for item in data:
                    item_year, item_month, _ = map(int, item.date.split('/'))
                    if item_month == month and item_year == year and item.mission_type != 'half':
                        normal_missions.append({'factory': item.factory, 'date': item.date, 'type': item.mission_type})
                normal_missions.sort(key=lambda x: x['date'])
                for mission in normal_missions:
                    filtered_data.append([mission['factory'], convert_to_persian_numbers(mission['date'])])
                    if mission['type'] == 'holiday':
                        filtered_data.append([f"{mission['factory']} (تعطیل)", convert_to_persian_numbers(mission['date'])])

                asalat_missions = []
                for item in data:
                    item_year, item_month, _ = map(int, item.date.split('/'))
                    if item_month == month and item_year == year and item.mission_type == 'half':
                        asalat_missions.append({'factory': item.factory, 'date': item.date})
                asalat_missions.sort(key=lambda x: x['date'])

                if asalat_missions:
                    filtered_data.append(["", "اصالت"])

                for mission in asalat_missions:
                    filtered_data.append([mission['factory'], convert_to_persian_numbers(mission['date'])])
            else:
                temp_data = []
                for item in data:
                    item_year, item_month, _ = map(int, item.date.split('/'))
                    if item_month == month and item_year == year:
                        if model_type == 'expense':
                            temp_data.append({'date': item.date, 'factory': item.factory, 'description': item.description, 'amount': item.amount})
                            total_amount += item.amount
                        elif model_type == 'khodro':
                            temp_data.append({'date': item.date, 'amount': item.amount, 'description': item.description, 'kilometer': item.kilometer})
                            total_amount += item.amount
                temp_data.sort(key=lambda x: x['date'])
                for index, item in enumerate(temp_data, 1):
                    if model_type == 'expense':
                        filtered_data.append([item['factory'], convert_to_persian_numbers(item['amount']), item['description'], convert_to_persian_numbers(item['date'])])
                    elif model_type == 'khodro':
                        filtered_data.append([convert_to_persian_numbers(item['amount']), item['description'], convert_to_persian_numbers(item['kilometer']), convert_to_persian_numbers(item['date'])])

            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{output_name}"'
            doc = SimpleDocTemplate(response, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
            pdfmetrics.registerFont(TTFont('Vazir', FONT_PATH))

            styles = getSampleStyleSheet()
            styles.add(ParagraphStyle(name='RTL', fontName='Vazir', fontSize=8, alignment=1, textColor=colors.black, leading=6))

            table_data = []
            rtl_headers = [Paragraph(get_display(reshape(header)), styles['RTL']) for header in headers]
            table_data.append(rtl_headers)

            for index, row in enumerate(filtered_data, 1):
                rtl_row = []
                for item in row:
                    if item == "اصالت":
                        asalat_style = ParagraphStyle(
                            'AsalatStyle',
                            parent=styles['RTL'],
                            fontSize=8,
                            textColor=colors.darkblue,
                            alignment=1,
                            fontName='Vazir'
                        )
                        rtl_row.append(Paragraph(get_display(reshape(str(item))), asalat_style))
                    else:
                        rtl_row.append(Paragraph(get_display(reshape(str(item))), styles['RTL']))
                rtl_row.append(Paragraph(convert_to_persian_numbers(index), styles['RTL']))
                table_data.append(rtl_row)

            if model_type == 'mission':
                col_widths = [150, 90, 40]
            elif model_type == 'expense':
                col_widths = [90, 90, 180, 90, 40]
            elif model_type == 'khodro':
                col_widths = [90, 180, 70, 90, 40]

            table = Table(table_data, colWidths=col_widths, repeatRows=1)
            table.setStyle(TableStyle([
                ('FONT', (0, 0), (-1, -1), 'Vazir'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('FONTSIZE', (0, 1), (-1, -1), 11),
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('LEFTPADDING', (0, 0), (-1, -1), 6),
                ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ]))

            for i, row in enumerate(table_data):
                if isinstance(row[0], Paragraph) and row[0].text == "اصالت":
                    table.setStyle(TableStyle([
                        ('BACKGROUND', (0, i), (-1, i), colors.darkgrey),
                        ('TEXTCOLOR', (0, i), (-1, i), colors.whitesmoke),  # رنگ متن روشن
                        ('FONTSIZE', (0, i), (-1, i), 14),      
                    ]))
                    break

            elements = []
            title_style = ParagraphStyle('CustomTitle', parent=styles['RTL'], fontSize=16, spaceAfter=30)
            title_text = get_display(reshape(f"{title} - {year}/{month:02d}"))
            elements.append(Paragraph(title_text, title_style))
            elements.append(table)

            if model_type in ['expense', 'khodro'] and filtered_data:
                total_style = ParagraphStyle('Total', parent=styles['RTL'], fontSize=10, spaceBefore=20)
                total_text = get_display(reshape(f"جمع کل: {convert_to_persian_numbers(total_amount)} ریال"))
                elements.append(Paragraph(total_text, total_style))

            elements.append(Paragraph("<br/><br/><br/>", styles['RTL']))

            signature_style = ParagraphStyle('Signature', parent=styles['RTL'], fontSize=12, spaceBefore=150)
            signature_table_data = [[
                Paragraph(get_display(reshape("مدیر عامل")), signature_style),
                Paragraph(get_display(reshape("کنترل کننده")), signature_style),
                Paragraph(get_display(reshape(f"{request.user.get_full_name()}")), signature_style)
            ]]
            signature_table = Table(signature_table_data, colWidths=[180, 180, 180])
            signature_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (0, 0), 'LEFT'),
                ('ALIGN', (1, 0), (1, 0), 'CENTER'),
                ('ALIGN', (2, 0), (2, 0), 'RIGHT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('LEFTPADDING', (0, 0), (-1, -1), 0),
                ('RIGHTPADDING', (0, 0), (-1, -1), 0),
                ('TOPPADDING', (0, 0), (-1, -1), 0),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
            ]))
            elements.append(signature_table)

            doc.build(elements)
            return response
    else:
        form = ReportForm()
    return render(request, 'report.html', {'form': form})



def convert_to_persian_numbers(text):
    """اعداد انگلیسی به فارسی + جداکننده سه‌رقمی"""
    english_numbers = "0123456789"
    persian_numbers = "۰۱۲۳۴۵۶۷۸۹"
    translation_table = str.maketrans(english_numbers, persian_numbers)
    
    text = str(text)
    if text.isdigit():
        text = "{:,}".format(int(text))
    return text.translate(translation_table)


def convert_index_to_persian(index):
    """شماره ردیف به فارسی بدون جداکننده"""
    english_numbers = "0123456789"
    persian_numbers = "۰۱۲۳۴۵۶۷۸۹"
    return str(index).translate(str.maketrans(english_numbers, persian_numbers))
